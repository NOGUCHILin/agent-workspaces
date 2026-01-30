"""正社員登用条件のPDFとスプレッドシートを生成するスクリプト"""

from pathlib import Path
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# 出力先
OUTPUT_DIR = Path(__file__).parent.parent / "documents"
OUTPUT_DIR.mkdir(exist_ok=True)

# 日本語フォント登録（CIDフォントを使用）
pdfmetrics.registerFont(UnicodeCIDFont('HeiseiKakuGo-W5'))


def create_pdf():
    """PDFを生成"""
    pdf_path = OUTPUT_DIR / "正社員登用条件_メフル.pdf"
    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=A4,
        leftMargin=20*mm,
        rightMargin=20*mm,
        topMargin=20*mm,
        bottomMargin=20*mm
    )

    # スタイル定義
    styles = {
        "title": ParagraphStyle(
            "title",
            fontName="HeiseiKakuGo-W5",
            fontSize=18,
            leading=24,
            alignment=1,  # center
            spaceAfter=10*mm
        ),
        "heading": ParagraphStyle(
            "heading",
            fontName="HeiseiKakuGo-W5",
            fontSize=14,
            leading=20,
            spaceBefore=8*mm,
            spaceAfter=4*mm
        ),
        "subheading": ParagraphStyle(
            "subheading",
            fontName="HeiseiKakuGo-W5",
            fontSize=12,
            leading=16,
            spaceBefore=4*mm,
            spaceAfter=2*mm
        ),
        "body": ParagraphStyle(
            "body",
            fontName="HeiseiKakuGo-W5",
            fontSize=11,
            leading=16,
            spaceBefore=2*mm
        ),
        "important": ParagraphStyle(
            "important",
            fontName="HeiseiKakuGo-W5",
            fontSize=11,
            leading=16,
            spaceBefore=2*mm,
            backColor=colors.Color(1, 1, 0.8)  # light yellow
        ),
        "input_field": ParagraphStyle(
            "input_field",
            fontName="HeiseiKakuGo-W5",
            fontSize=12,
            leading=20,
            spaceBefore=4*mm,
            spaceAfter=4*mm
        ),
    }

    # コンテンツ作成
    content = []

    # タイトル
    content.append(Paragraph("正社員登用の条件", styles["title"]))

    # 1. 期間
    content.append(Paragraph("1. 期間", styles["heading"]))
    content.append(Paragraph("VISA延長の関係で、正社員登用の判断までの期限：", styles["body"]))
    content.append(Spacer(1, 5*mm))
    content.append(Paragraph("→ ＿＿＿＿年＿＿月＿＿日まで（記入してください）", styles["input_field"]))

    # 2. 必須スキル
    content.append(Paragraph("2. 必須スキル", styles["heading"]))

    content.append(Paragraph("【最重要（必須）】", styles["subheading"]))
    content.append(Paragraph("● 現場リーダーとしての状況確認・指示出し", styles["important"]))
    content.append(Paragraph("　- 効率よく業務を進められるよう全体を見て判断・指示できること", styles["body"]))
    content.append(Paragraph("　- これが一番重要かつ習得が難しい", styles["body"]))

    content.append(Paragraph("【基本スキル（期限内に習得目標）】", styles["subheading"]))
    content.append(Paragraph("● 開封", styles["body"]))
    content.append(Paragraph("● 査定", styles["body"]))
    content.append(Paragraph("● アクティベート", styles["body"]))

    content.append(Paragraph("【習得は要相談（期限内に難しければ応相談）】", styles["subheading"]))
    content.append(Paragraph("● 出品　● 検品　● 発送準備・送り状作成", styles["body"]))
    content.append(Paragraph("● 成約仕分　● BM在庫確認　● 電話対応", styles["body"]))

    # 3. 勤務態度
    content.append(Paragraph("3. 勤務態度", styles["heading"]))
    content.append(Paragraph("● 遅刻・欠勤ともに月1回以下", styles["body"]))
    content.append(Paragraph("● 効率よく業務を回すために常に頭を回転させる姿勢", styles["body"]))

    # 4. マインドセット
    content.append(Paragraph("4. マインドセット（重要）", styles["heading"]))
    content.append(Paragraph(
        "弊社はスタッフ数が少ない分、正社員の重みが他の企業とは段違いに大きいです。",
        styles["important"]
    ))
    content.append(Spacer(1, 2*mm))
    content.append(Paragraph(
        "正社員になるなら運命共同体レベルで働く覚悟が必要です。",
        styles["important"]
    ))
    content.append(Spacer(1, 2*mm))
    content.append(Paragraph(
        "この認識をしっかり持った上で判断してください。",
        styles["body"]
    ))

    doc.build(content)
    print(f"PDF作成完了: {pdf_path}")
    return pdf_path


def create_excel():
    """スプレッドシートを生成"""
    xlsx_path = OUTPUT_DIR / "正社員登用条件_メフル.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "正社員登用条件"

    # スタイル定義
    title_font = Font(name="ヒラギノ角ゴシック W3", size=16, bold=True)
    heading_font = Font(name="ヒラギノ角ゴシック W3", size=12, bold=True)
    body_font = Font(name="ヒラギノ角ゴシック W3", size=11)
    important_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 列幅設定
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15

    row = 1

    # タイトル
    ws.merge_cells('A1:D1')
    ws['A1'] = "正社員登用の条件"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    row += 2

    # 1. 期間
    ws[f'A{row}'] = "1."
    ws[f'B{row}'] = "期間"
    ws[f'B{row}'].font = heading_font
    row += 1
    ws[f'B{row}'] = "VISA延長の関係で、正社員登用の判断までの期限："
    ws[f'B{row}'].font = body_font
    row += 1
    ws[f'B{row}'] = "期限日"
    ws[f'C{row}'] = ""  # 入力欄
    ws[f'C{row}'].border = thin_border
    ws[f'C{row}'].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    ws[f'D{row}'] = "（記入してください）"
    ws[f'D{row}'].font = Font(name="ヒラギノ角ゴシック W3", size=10, color="666666")
    row += 2

    # 2. 必須スキル
    ws[f'A{row}'] = "2."
    ws[f'B{row}'] = "必須スキル"
    ws[f'B{row}'].font = heading_font
    row += 1

    # 最重要
    ws[f'B{row}'] = "【最重要（必須）】"
    ws[f'B{row}'].font = Font(name="ヒラギノ角ゴシック W3", size=11, bold=True)
    row += 1
    ws[f'B{row}'] = "現場リーダーとしての状況確認・指示出し"
    ws[f'B{row}'].font = body_font
    ws[f'B{row}'].fill = important_fill
    ws[f'C{row}'] = "効率よく業務を進められるよう全体を見て判断・指示できること。これが一番重要かつ習得が難しい。"
    ws[f'C{row}'].font = body_font
    ws[f'C{row}'].fill = important_fill
    row += 1

    # 基本スキル
    ws[f'B{row}'] = "【基本スキル（期限内に習得目標）】"
    ws[f'B{row}'].font = Font(name="ヒラギノ角ゴシック W3", size=11, bold=True)
    row += 1
    for skill in ["開封", "査定", "アクティベート"]:
        ws[f'B{row}'] = f"● {skill}"
        ws[f'B{row}'].font = body_font
        row += 1

    # 要相談スキル
    ws[f'B{row}'] = "【習得は要相談（期限内に難しければ応相談）】"
    ws[f'B{row}'].font = Font(name="ヒラギノ角ゴシック W3", size=11, bold=True)
    row += 1
    for skill in ["出品", "検品", "発送準備・送り状作成", "成約仕分", "BM在庫確認", "電話対応"]:
        ws[f'B{row}'] = f"● {skill}"
        ws[f'B{row}'].font = body_font
        row += 1

    row += 1

    # 3. 勤務態度
    ws[f'A{row}'] = "3."
    ws[f'B{row}'] = "勤務態度"
    ws[f'B{row}'].font = heading_font
    row += 1
    ws[f'B{row}'] = "● 遅刻・欠勤ともに月1回以下"
    ws[f'B{row}'].font = body_font
    row += 1
    ws[f'B{row}'] = "● 効率よく業務を回すために常に頭を回転させる姿勢"
    ws[f'B{row}'].font = body_font
    row += 2

    # 4. マインドセット
    ws[f'A{row}'] = "4."
    ws[f'B{row}'] = "マインドセット（重要）"
    ws[f'B{row}'].font = heading_font
    row += 1

    messages = [
        "弊社はスタッフ数が少ない分、正社員の重みが他の企業とは段違いに大きいです。",
        "正社員になるなら運命共同体レベルで働く覚悟が必要です。",
        "この認識をしっかり持った上で判断してください。"
    ]
    for msg in messages:
        ws.merge_cells(f'B{row}:C{row}')
        ws[f'B{row}'] = msg
        ws[f'B{row}'].font = body_font
        if "段違い" in msg or "運命共同体" in msg:
            ws[f'B{row}'].fill = important_fill
        row += 1

    wb.save(xlsx_path)
    print(f"Excel作成完了: {xlsx_path}")
    return xlsx_path


if __name__ == "__main__":
    create_pdf()
    create_excel()
