import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# データ読み込み
buyback_df = pd.read_csv('買取価格_filtered.csv')
sales_df = pd.read_csv('販売価格_filtered.csv')

print("=== データ読み込み完了 ===")
print(f"買取価格: {len(buyback_df)}行")
print(f"販売価格: {len(sales_df)}行\n")

# 正規化: 機種名の統一
print("=== 正規化処理 ===")
buyback_df['機体型番'] = buyback_df['機体型番'].replace({
    'iPhone SE（第1世代）': 'iPhone SE1',
    'iPhone SE（第2世代）': 'iPhone SE2',
    'iPhone SE（第3世代）': 'iPhone SE3'
})
print("iPhone SE表記を統一しました")

# iPhone 7, iPhone 8を除外
before_count = len(buyback_df)
buyback_df = buyback_df[~buyback_df['機体型番'].isin(['iPhone 7', 'iPhone 8', 'iPhone 8 Plus'])].copy()
after_count = len(buyback_df)
print(f"iPhone 7/8を除外: {before_count}行 → {after_count}行\n")

# 販売価格データをピボット化（機種・容量・グレードで検索しやすくする）
sales_pivot = sales_df.pivot_table(
    index=['機種', '容量'],
    columns='グレード',
    values='平均売価',
    aggfunc='first'
).reset_index()

print("=== ランク・グレード対応表 ===")
print("新品同様 → A")
print("美品 → (A + B) / 2")
print("使用感あり → (B + C) / 2")
print("目立つ傷あり → C\n")

# 買取価格データに販売価格を紐付け
result_rows = []

for _, row in buyback_df.iterrows():
    model = row['機体型番']
    capacity = row['記憶容量']
    rank = row['等級']
    high_price = row['高額買取価格']
    express_price = row['特急買取価格']

    # 販売価格データを検索
    sales_data = sales_pivot[
        (sales_pivot['機種'] == model) &
        (sales_pivot['容量'] == capacity)
    ]

    if len(sales_data) == 0:
        print(f"⚠ 販売価格が見つかりません: {model} {capacity}")
        continue

    sales_data = sales_data.iloc[0]

    # ランクに応じた販売価格を計算
    if rank == '新品同様':
        sales_price = sales_data['A']
    elif rank == '美品':
        sales_price = (sales_data['A'] + sales_data['B']) / 2
    elif rank == '使用感あり':
        sales_price = (sales_data['B'] + sales_data['C']) / 2
    elif rank == '目立つ傷あり':
        sales_price = sales_data['C']
    else:
        print(f"⚠ 不明なランク: {rank}")
        continue

    # 手数料引き粗利計算（販売価格 × 0.89 - 買取価格）
    sales_after_fee = sales_price * 0.89
    profit_high = sales_after_fee - high_price
    profit_express = sales_after_fee - express_price

    result_rows.append({
        '機種': model,
        '容量': capacity,
        'ランク': rank,
        '高額買取価格': high_price,
        '特急買取価格': express_price,
        '販売価格': sales_price,
        '手数料引粗利（高額）': profit_high,
        '手数料引粗利（特急）': profit_express,
        '新高額買取価格': '',
        '新特急買取価格': '',
        '新手数料引粗利（高額）': '',
        '新手数料引粗利（特急）': ''
    })

result_df = pd.DataFrame(result_rows)

print(f"\n=== 結果 ===")
print(f"紐付け成功: {len(result_df)}行\n")

# Excelファイルに書き出し（関数を含める）
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "粗利分析"

# ヘッダー行
headers = ['機種', '容量', 'ランク', '高額買取価格', '特急買取価格', '販売価格',
           '手数料引粗利（高額）', '手数料引粗利（特急）',
           '新高額買取価格', '新特急買取価格',
           '新手数料引粗利（高額）', '新手数料引粗利（特急）']

# ヘッダーを書き込み
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    cell.alignment = Alignment(horizontal='center')

# データを書き込み
for row_idx, row_data in enumerate(result_df.itertuples(index=False), start=2):
    ws.cell(row=row_idx, column=1, value=row_data[0])  # 機種
    ws.cell(row=row_idx, column=2, value=row_data[1])  # 容量
    ws.cell(row=row_idx, column=3, value=row_data[2])  # ランク
    ws.cell(row=row_idx, column=4, value=row_data[3])  # 高額買取価格
    ws.cell(row=row_idx, column=5, value=row_data[4])  # 特急買取価格
    ws.cell(row=row_idx, column=6, value=row_data[5])  # 販売価格
    ws.cell(row=row_idx, column=7, value=row_data[6])  # 手数料引粗利（高額）
    ws.cell(row=row_idx, column=8, value=row_data[7])  # 手数料引粗利（特急）
    # 新高額買取価格（入力欄）
    # 新特急買取価格（入力欄）
    # 新手数料引粗利（高額）= F * 0.89 - I
    ws.cell(row=row_idx, column=11, value=f"=F{row_idx}*0.89-I{row_idx}")
    # 新手数料引粗利（特急）= F * 0.89 - J
    ws.cell(row=row_idx, column=12, value=f"=F{row_idx}*0.89-J{row_idx}")

# 列幅を調整
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 18
ws.column_dimensions['H'].width = 18
ws.column_dimensions['I'].width = 18
ws.column_dimensions['J'].width = 18
ws.column_dimensions['K'].width = 22
ws.column_dimensions['L'].width = 22

# 保存
output_file = '粗利分析_20251118.xlsx'
wb.save(output_file)

print(f"✓ 保存完了: {output_file}")
print(f"  - 全{len(result_df)}行")
print(f"  - 新高額買取価格・新特急買取価格に値を入力すると、自動で新粗利が計算されます")
