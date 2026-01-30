"""
買取価格最適化ファイル生成スクリプト

使い方:
    uv run python generate_optimizer.py

機能:
    - 最新の買取価格・販売価格CSVを自動検出
    - BM平均売価シートから背景色を取得し販売価格列に反映
    - 買取価格最適化Excelを生成（日付付き）
    - 条件付き書式（G/H列: 1万未満は赤文字）
    - I/J列にデフォルト数式（=D-5000, =E-5000）
    - 古いファイルをarchiveフォルダに移動
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import CellIsRule
from pathlib import Path
from datetime import datetime
import glob
import shutil
import urllib.request
import tempfile

# 除外する機種
EXCLUDE_MODELS = ['iPhone 7', 'iPhone 8', 'iPhone 8 Plus']

# 機種名の正規化マッピング
MODEL_NORMALIZATION = {
    'iPhone SE（第1世代）': 'iPhone SE1',
    'iPhone SE（第2世代）': 'iPhone SE2',
    'iPhone SE（第3世代）': 'iPhone SE3'
}

# ランク→BMグレードのマッピング（色検索用）
RANK_TO_GRADES = {
    '新品同様': ['A'],
    '美品': ['A', 'B'],
    '使用感あり': ['B', 'C'],
    '目立つ傷あり': ['C'],
}

# BM平均売価シートのエクスポートURL（gid=970596964）
BM_EXPORT_URL = "https://docs.google.com/spreadsheets/d/1Gg4Lvvlx25GGk-LdEnr8apUO2Q4e2ZOYovaAlBfV7os/export?format=xlsx&gid=970596964"

# ============================================================
# BM色取得
# ============================================================

def build_bm_color_lookup():
    """BM平均売価シートからセル背景色のルックアップテーブルを構築"""
    print("  BM平均売価シートをダウンロード中...")

    # ローカルキャッシュを確認
    cache_path = Path("bm_avg_price_temp.xlsx")

    # 常に最新をダウンロード
    try:
        urllib.request.urlretrieve(BM_EXPORT_URL, str(cache_path))
        print("  ダウンロード完了")
    except Exception as e:
        if cache_path.exists():
            print(f"  ダウンロード失敗（キャッシュ使用）: {e}")
        else:
            print(f"  ⚠ BM色データ取得失敗: {e}")
            return {}

    bm_wb = openpyxl.load_workbook(str(cache_path), data_only=True)
    bm_ws = bm_wb.active

    color_lookup = {}
    for row in range(2, bm_ws.max_row + 1):
        model = bm_ws.cell(row=row, column=1).value
        capacity = bm_ws.cell(row=row, column=2).value
        grade = bm_ws.cell(row=row, column=3).value

        if not model or not capacity or not grade:
            continue

        cell = bm_ws.cell(row=row, column=4)  # 平均売価列
        fill = cell.fill
        color_hex = None
        if fill.fill_type == 'solid' and fill.fgColor and fill.fgColor.rgb:
            rgb = str(fill.fgColor.rgb)
            if rgb not in ('FFFFFFFF', '00000000', '0'):
                color_hex = rgb

        color_lookup[(str(model), str(capacity), str(grade))] = color_hex

    bm_wb.close()
    print(f"  色ルックアップ: {len(color_lookup)}エントリ")
    return color_lookup

def get_bm_color_for_row(color_lookup, model, capacity, rank):
    """指定のモデル・容量・ランクに対応するBM背景色を取得"""
    grades = RANK_TO_GRADES.get(rank, [])
    for grade in grades:
        key = (str(model), str(capacity), grade)
        if key in color_lookup and color_lookup[key]:
            return color_lookup[key]
    return None

# ============================================================
# メイン処理
# ============================================================

def find_latest_csv(pattern):
    """最新のCSVファイルを検索"""
    files = glob.glob(pattern)
    if not files:
        raise FileNotFoundError(f"ファイルが見つかりません: {pattern}")
    # 日付でソートして最新を取得
    files.sort(reverse=True)
    return files[0]

def archive_old_files(pattern, archive_dir):
    """古いファイルをアーカイブフォルダに移動"""
    archive_path = Path(archive_dir)
    archive_path.mkdir(exist_ok=True)

    files = glob.glob(pattern)
    for file in files:
        file_path = Path(file)
        dest = archive_path / file_path.name
        # 同名ファイルがあれば上書き
        if dest.exists():
            dest.unlink()
        shutil.move(file, dest)
        print(f"  アーカイブ: {file_path.name} -> archive/")

def main():
    print("=" * 60)
    print("買取価格最適化ファイル生成")
    print("=" * 60)

    # 1. 最新のCSVファイルを検索
    print("\n[1] 最新CSVファイルを検索中...")
    buyback_file = find_latest_csv("price-data/買取価格[0-9]*.csv")
    sales_file = find_latest_csv("price-data/販売価格[0-9]*.csv")

    print(f"  買取価格: {buyback_file}")
    print(f"  販売価格: {sales_file}")

    # 2. データ読み込み
    print("\n[2] データ読み込み中...")
    buyback_df = pd.read_csv(buyback_file)
    sales_df = pd.read_csv(sales_file)

    print(f"  買取価格: {len(buyback_df)}行")
    print(f"  販売価格: {len(sales_df)}行")

    # 3. データ正規化・フィルタリング
    print("\n[3] データ正規化・フィルタリング中...")

    # 買取価格から不要な等級を除外
    buyback_df = buyback_df[~buyback_df['等級'].isin(['新品・未開封', '外装ジャンク'])].copy()

    # 販売価格からプレミアムを除外
    sales_df = sales_df[sales_df['グレード'] != 'プレミアム'].copy()

    # 販売価格を数値に変換（#DIV/0!などのエラー値を除外）
    sales_df['平均売価'] = pd.to_numeric(sales_df['平均売価'], errors='coerce')
    sales_df = sales_df.dropna(subset=['平均売価'])

    # 機種名の正規化
    buyback_df['機体型番'] = buyback_df['機体型番'].replace(MODEL_NORMALIZATION)

    # 除外機種を削除
    before_count = len(buyback_df)
    buyback_df = buyback_df[~buyback_df['機体型番'].isin(EXCLUDE_MODELS)].copy()
    after_count = len(buyback_df)

    print(f"  正規化完了: {before_count}行 -> {after_count}行")

    # 4. 販売価格データをピボット化
    sales_pivot = sales_df.pivot_table(
        index=['機種', '容量'],
        columns='グレード',
        values='平均売価',
        aggfunc='first'
    ).reset_index()

    # 5. 買取価格と販売価格を紐付け
    print("\n[4] 買取価格と販売価格を紐付け中...")

    result_rows = []

    for _, row in buyback_df.iterrows():
        model = row['機体型番']
        capacity = row['記憶容量']
        rank = row['等級']
        high_price = row['高額買取価格']
        express_price = row['特急買取価格']

        # 販売価格データを検索
        sales_data = sales_pivot[
            (sales_pivot['機種'] == model) &
            (sales_pivot['容量'] == capacity)
        ]

        if len(sales_data) == 0:
            continue

        sales_data = sales_data.iloc[0]

        # ランクに応じた販売価格を計算
        if rank == '新品同様':
            sales_price = sales_data['A']
        elif rank == '美品':
            sales_price = (sales_data['A'] + sales_data['B']) / 2
        elif rank == '使用感あり':
            sales_price = (sales_data['B'] + sales_data['C']) / 2
        elif rank == '目立つ傷あり':
            sales_price = sales_data['C']
        else:
            continue

        # 手数料引き後の販売価格
        sales_after_fee = sales_price * 0.89

        # 手数料引き粗利計算
        profit_high = sales_after_fee - high_price
        profit_express = sales_after_fee - express_price

        result_rows.append({
            '機種': model,
            '容量': capacity,
            'ランク': rank,
            '高額買取': high_price,
            '特急買取': express_price,
            '販売価格': sales_price,
            '手引粗(高額)': profit_high,
            '手引粗(特急)': profit_express,
            '新高額買取': '',
            '新特急買取': '',
            '新手引粗(高額)': '',
            '新手引粗(特急)': '',
        })

    result_df = pd.DataFrame(result_rows)

    print(f"  紐付け成功: {len(result_df)}行")

    # 6. BM平均売価の背景色を取得
    print("\n[5] BM平均売価の背景色を取得中...")
    color_lookup = build_bm_color_lookup()

    # 7. Excelファイルに書き出し
    print("\n[6] Excelファイル作成中...")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "買取価格最適化"

    # ヘッダー行（12列）
    headers = ['機種', '容量', 'ランク', '高額買取', '特急買取', '販売価格',
               '手引粗(高額)', '手引粗(特急)',
               '新高額買取', '新特急買取', '新手引粗(高額)', '新手引粗(特急)']

    # BM平均売価シートの色（販売価格列ヘッダー用）
    bm_header_fill = PatternFill(start_color="666666", end_color="666666", fill_type="solid")
    bm_header_font = Font(bold=True, size=11, color="F3F3F3")

    # ヘッダーを書き込み
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        if col_idx == 6:  # 販売価格列: BM平均売価の色
            cell.font = bm_header_font
            cell.fill = bm_header_fill
        else:
            cell.font = Font(bold=True, size=11, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # データを書き込み
    for row_idx, row_data in enumerate(result_df.itertuples(index=False), start=2):
        model = row_data[0]
        capacity = row_data[1]
        rank = row_data[2]

        ws.cell(row=row_idx, column=1, value=model)       # 機種
        ws.cell(row=row_idx, column=2, value=capacity)     # 容量
        ws.cell(row=row_idx, column=3, value=rank)         # ランク
        ws.cell(row=row_idx, column=4, value=row_data[3])  # 高額買取
        ws.cell(row=row_idx, column=5, value=row_data[4])  # 特急買取

        # 販売価格（BM平均売価の動的背景色）
        cell_f = ws.cell(row=row_idx, column=6, value=row_data[5])
        bm_color = get_bm_color_for_row(color_lookup, model, capacity, rank)
        if bm_color:
            hex_color = bm_color[2:] if len(bm_color) == 8 else bm_color
            cell_f.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

        ws.cell(row=row_idx, column=7, value=row_data[6])  # 手引粗(高額)
        ws.cell(row=row_idx, column=8, value=row_data[7])  # 手引粗(特急)

        # 新高額買取（I列: デフォルト=D-5000）
        ws.cell(row=row_idx, column=9, value=f"=D{row_idx}-5000")
        # 新特急買取（J列: デフォルト=E-5000）
        ws.cell(row=row_idx, column=10, value=f"=E{row_idx}-5000")
        # 新手引粗(高額) = F * 0.89 - I
        ws.cell(row=row_idx, column=11, value=f"=F{row_idx}*0.89-I{row_idx}")
        # 新手引粗(特急) = F * 0.89 - J
        ws.cell(row=row_idx, column=12, value=f"=F{row_idx}*0.89-J{row_idx}")

    # 条件付き書式: G列・H列で1万未満は赤文字
    data_rows = len(result_df) + 1
    red_font = Font(color="FF0000")
    ws.conditional_formatting.add(
        f"G2:G{data_rows}",
        CellIsRule(operator='lessThan', formula=['10000'], font=red_font)
    )
    ws.conditional_formatting.add(
        f"H2:H{data_rows}",
        CellIsRule(operator='lessThan', formula=['10000'], font=red_font)
    )

    # 列幅を調整
    column_widths = {
        'A': 22, 'B': 10, 'C': 14, 'D': 12, 'E': 12, 'F': 12,
        'G': 14, 'H': 14, 'I': 13, 'J': 13, 'K': 16, 'L': 16
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 数値フォーマット
    for row in ws.iter_rows(min_row=2, max_row=len(result_df)+1, min_col=4, max_col=8):
        for cell in row:
            cell.number_format = '#,##0'

    for row in ws.iter_rows(min_row=2, max_row=len(result_df)+1, min_col=11, max_col=12):
        for cell in row:
            cell.number_format = '#,##0'

    # 8. ファイル保存
    today = datetime.now().strftime("%Y%m%d")
    output_file = f'買取価格最適化_{today}.xlsx'
    wb.save(output_file)

    print(f"  ✓ 保存完了: {output_file}")
    print(f"  - メインシート: {len(result_df)}行")

    # 9. 古いファイルをアーカイブ
    print("\n[7] 古いファイルをアーカイブ中...")
    archive_old_files('買取価格最適化_*.xlsx', 'archive')
    # 最新ファイルをアーカイブから戻す
    if Path(f'archive/{output_file}').exists():
        shutil.move(f'archive/{output_file}', output_file)
        print(f"  最新ファイルを戻しました: {output_file}")

    print("\n" + "=" * 60)
    print("✅ 完了")
    print("=" * 60)
    print(f"\n生成されたファイル: {output_file}")
    print("使い方: I列・J列に新価格を入力すると、自動で粗利が計算されます")

if __name__ == '__main__':
    main()
