"""
効果分析レポート生成スクリプト

使い方:
    # 特定の価格変更日を指定
    uv run python scripts/generate_report.py --change-date 2025-11-18

    # すべての価格変更のレポート生成
    uv run python scripts/generate_report.py --all

機能:
    - 価格変更効果をExcelレポートとして出力
    - 機種・容量・ランク別の詳細分析
    - グラフによる可視化
    - 色分けによる判定（改善/悪化/変化なし）
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
import argparse
from pathlib import Path
from datetime import datetime, timedelta

# ============================================================
# 設定
# ============================================================

BASE_DIR = Path(__file__).parent.parent
PRICE_CHANGES_DIR = BASE_DIR / "data" / "price_changes"
RESULTS_DIR = BASE_DIR / "data" / "results"

BEFORE_DAYS = 7
AFTER_DAYS = 7

# 色設定
COLOR_GREEN = "C6EFCE"   # 改善
COLOR_YELLOW = "FFEB9C"  # 変化なし
COLOR_RED = "FFC7CE"     # 悪化
COLOR_HEADER = "4472C4"  # ヘッダー

# ============================================================
# データ読み込み
# ============================================================

def load_collected_data() -> pd.DataFrame:
    """最新の収集済みデータを読み込み"""
    csv_files = list(RESULTS_DIR.glob("collected_data_*.csv"))
    if not csv_files:
        print("❌ エラー: 収集済みデータが見つかりません")
        return pd.DataFrame()

    latest_file = max(csv_files, key=lambda f: f.stat().st_mtime)
    df = pd.read_csv(latest_file)
    df['date'] = pd.to_datetime(df['date'])
    return df


def load_price_changes() -> pd.DataFrame:
    """価格変更履歴を読み込み"""
    price_changes_file = PRICE_CHANGES_DIR / "price_changes.csv"
    if not price_changes_file.exists():
        print(f"❌ エラー: {price_changes_file} が見つかりません")
        return pd.DataFrame()

    df = pd.read_csv(price_changes_file)
    df['変更日'] = pd.to_datetime(df['変更日'])
    return df


# ============================================================
# レポート生成
# ============================================================

def generate_report(change_date: datetime, df_data: pd.DataFrame, df_price_changes: pd.DataFrame):
    """
    価格変更効果レポートをExcel形式で生成

    Args:
        change_date: 価格変更日
        df_data: 収集済みデータ
        df_price_changes: 価格変更履歴
    """
    print(f"📊 レポート生成中: {change_date.strftime('%Y-%m-%d')}")

    # 期間設定
    before_start = change_date - timedelta(days=BEFORE_DAYS)
    before_end = change_date - timedelta(days=1)
    after_start = change_date
    after_end = change_date + timedelta(days=AFTER_DAYS - 1)

    # 対象の価格変更を抽出
    df_changes = df_price_changes[df_price_changes['変更日'] == change_date].copy()

    if df_changes.empty:
        print(f"⚠️  警告: {change_date.strftime('%Y-%m-%d')} の価格変更が見つかりません")

    # Excelファイル作成
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除

    # ============================================================
    # シート1: サマリー
    # ============================================================
    ws_summary = wb.create_sheet("サマリー")

    # ヘッダー情報
    ws_summary['A1'] = '買取価格変更 効果分析レポート'
    ws_summary['A1'].font = Font(size=16, bold=True)

    ws_summary['A3'] = '変更日:'
    ws_summary['B3'] = change_date.strftime('%Y-%m-%d')

    ws_summary['A4'] = '分析期間:'
    ws_summary['B4'] = f'変更前: {before_start.strftime("%Y-%m-%d")} ～ {before_end.strftime("%Y-%m-%d")}'
    ws_summary['B5'] = f'変更後: {after_start.strftime("%Y-%m-%d")} ～ {after_end.strftime("%Y-%m-%d")}'

    # 変更前・変更後データを抽出
    df_before = df_data[
        (df_data['date'] >= before_start) & (df_data['date'] <= before_end)
    ]
    df_after = df_data[
        (df_data['date'] >= after_start) & (df_data['date'] <= after_end)
    ]

    # 全体サマリー
    before_total_estimates = df_before['count_estimate'].sum()
    before_total_kits = df_before['count_kit'].sum()
    before_cv_rate = (before_total_kits / before_total_estimates * 100) if before_total_estimates > 0 else 0

    after_total_estimates = df_after['count_estimate'].sum()
    after_total_kits = df_after['count_kit'].sum()
    after_cv_rate = (after_total_kits / after_total_estimates * 100) if after_total_estimates > 0 else 0

    ws_summary['A7'] = '【全体サマリー】'
    ws_summary['A7'].font = Font(bold=True, size=12)

    headers = ['', '変更前', '変更後', '変化']
    for col, header in enumerate(headers, start=1):
        cell = ws_summary.cell(row=8, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')

    rows_data = [
        ['仮査定数（週合計）', before_total_estimates, after_total_estimates,
         f"{((after_total_estimates - before_total_estimates) / before_total_estimates * 100) if before_total_estimates > 0 else 0:.1f}%"],
        ['梱包キット数（週合計）', before_total_kits, after_total_kits,
         f"{((after_total_kits - before_total_kits) / before_total_kits * 100) if before_total_kits > 0 else 0:.1f}%"],
        ['コンバージョン率', f"{before_cv_rate:.2f}%", f"{after_cv_rate:.2f}%",
         f"{after_cv_rate - before_cv_rate:+.2f}pt"],
    ]

    for row_idx, row_data in enumerate(rows_data, start=9):
        for col_idx, value in enumerate(row_data, start=1):
            ws_summary.cell(row=row_idx, column=col_idx, value=value)

    # ============================================================
    # シート2: 機種別詳細
    # ============================================================
    ws_detail = wb.create_sheet("機種別詳細")

    # ヘッダー
    detail_headers = [
        '機種', '容量', 'ランク',
        '変更前_仮査定', '変更前_キット', '変更前_CV率(%)',
        '変更後_仮査定', '変更後_キット', '変更後_CV率(%)',
        '仮査定_変化(%)', 'キット_変化(%)', 'CV率_変化(pt)', '判定'
    ]

    for col, header in enumerate(detail_headers, start=1):
        cell = ws_detail.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 機種・容量・ランク別の集計
    grouped_before = df_before.groupby(['model', 'capacity', 'rank']).agg({
        'count_estimate': 'sum',
        'count_kit': 'sum'
    }).reset_index()

    grouped_after = df_after.groupby(['model', 'capacity', 'rank']).agg({
        'count_estimate': 'sum',
        'count_kit': 'sum'
    }).reset_index()

    # マージ
    df_merged = pd.merge(
        grouped_before,
        grouped_after,
        on=['model', 'capacity', 'rank'],
        how='outer',
        suffixes=('_before', '_after')
    ).fillna(0)

    # CV率計算
    df_merged['cv_before'] = (df_merged['count_kit_before'] / df_merged['count_estimate_before'] * 100).fillna(0)
    df_merged['cv_after'] = (df_merged['count_kit_after'] / df_merged['count_estimate_after'] * 100).fillna(0)

    # 変化率計算
    df_merged['change_estimate'] = ((df_merged['count_estimate_after'] - df_merged['count_estimate_before']) /
                                     df_merged['count_estimate_before'] * 100).replace([float('inf'), -float('inf')], 0).fillna(0)

    df_merged['change_kit'] = ((df_merged['count_kit_after'] - df_merged['count_kit_before']) /
                                df_merged['count_kit_before'] * 100).replace([float('inf'), -float('inf')], 0).fillna(0)

    df_merged['change_cv'] = df_merged['cv_after'] - df_merged['cv_before']

    # 判定
    def judge(cv_change):
        if cv_change > 2:
            return '改善'
        elif cv_change < -2:
            return '悪化'
        else:
            return '変化なし'

    df_merged['判定'] = df_merged['change_cv'].apply(judge)

    # データを書き込み
    for row_idx, row in enumerate(df_merged.itertuples(), start=2):
        ws_detail.cell(row=row_idx, column=1, value=row.model)
        ws_detail.cell(row=row_idx, column=2, value=row.capacity)
        ws_detail.cell(row=row_idx, column=3, value=row.rank)
        ws_detail.cell(row=row_idx, column=4, value=int(row.count_estimate_before))
        ws_detail.cell(row=row_idx, column=5, value=int(row.count_kit_before))
        ws_detail.cell(row=row_idx, column=6, value=round(row.cv_before, 2))
        ws_detail.cell(row=row_idx, column=7, value=int(row.count_estimate_after))
        ws_detail.cell(row=row_idx, column=8, value=int(row.count_kit_after))
        ws_detail.cell(row=row_idx, column=9, value=round(row.cv_after, 2))
        ws_detail.cell(row=row_idx, column=10, value=round(row.change_estimate, 1))
        ws_detail.cell(row=row_idx, column=11, value=round(row.change_kit, 1))
        ws_detail.cell(row=row_idx, column=12, value=round(row.change_cv, 2))
        ws_detail.cell(row=row_idx, column=13, value=row.判定)

        # 判定に基づく色分け
        if row.判定 == '改善':
            fill_color = COLOR_GREEN
        elif row.判定 == '悪化':
            fill_color = COLOR_RED
        else:
            fill_color = COLOR_YELLOW

        ws_detail.cell(row=row_idx, column=13).fill = PatternFill(
            start_color=fill_color, end_color=fill_color, fill_type='solid'
        )

    # 列幅調整
    ws_detail.column_dimensions['A'].width = 20
    ws_detail.column_dimensions['B'].width = 12
    ws_detail.column_dimensions['C'].width = 15

    # ============================================================
    # シート3: 日次推移
    # ============================================================
    ws_daily = wb.create_sheet("日次推移")

    # 日次集計
    df_daily_before = df_before.groupby('date').agg({
        'count_estimate': 'sum',
        'count_kit': 'sum'
    }).reset_index()
    df_daily_before['cv_rate'] = (df_daily_before['count_kit'] / df_daily_before['count_estimate'] * 100).fillna(0)
    df_daily_before['period'] = '変更前'

    df_daily_after = df_after.groupby('date').agg({
        'count_estimate': 'sum',
        'count_kit': 'sum'
    }).reset_index()
    df_daily_after['cv_rate'] = (df_daily_after['count_kit'] / df_daily_after['count_estimate'] * 100).fillna(0)
    df_daily_after['period'] = '変更後'

    df_daily_combined = pd.concat([df_daily_before, df_daily_after]).sort_values('date')

    # ヘッダー
    daily_headers = ['日付', '期間', '仮査定数', '梱包キット数', 'CV率(%)']
    for col, header in enumerate(daily_headers, start=1):
        cell = ws_daily.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')

    # データ書き込み
    for row_idx, row in enumerate(df_daily_combined.itertuples(), start=2):
        ws_daily.cell(row=row_idx, column=1, value=row.date.strftime('%Y-%m-%d'))
        ws_daily.cell(row=row_idx, column=2, value=row.period)
        ws_daily.cell(row=row_idx, column=3, value=int(row.count_estimate))
        ws_daily.cell(row=row_idx, column=4, value=int(row.count_kit))
        ws_daily.cell(row=row_idx, column=5, value=round(row.cv_rate, 2))

    # ============================================================
    # 保存
    # ============================================================
    output_file = RESULTS_DIR / f"impact_report_{change_date.strftime('%Y%m%d')}.xlsx"
    wb.save(output_file)

    print(f"✅ レポート生成完了: {output_file}")
    print()


# ============================================================
# メイン処理
# ============================================================

def main():
    parser = argparse.ArgumentParser(description='買取価格変更効果レポート生成')
    parser.add_argument('--change-date', type=str, help='価格変更日 (YYYY-MM-DD)')
    parser.add_argument('--all', action='store_true', help='すべての価格変更のレポート生成')

    args = parser.parse_args()

    # データ読み込み
    df_data = load_collected_data()
    df_price_changes = load_price_changes()

    if df_data.empty or df_price_changes.empty:
        return

    if args.all:
        # すべての価格変更日のレポート生成
        change_dates = df_price_changes['変更日'].unique()
        print(f"📊 {len(change_dates)} 件のレポートを生成します\n")

        for change_date in sorted(change_dates):
            generate_report(pd.to_datetime(change_date), df_data, df_price_changes)

    elif args.change_date:
        # 特定日のレポート生成
        change_date = pd.to_datetime(args.change_date)
        generate_report(change_date, df_data, df_price_changes)

    else:
        print("❌ エラー: --change-date または --all を指定してください")
        parser.print_help()


if __name__ == "__main__":
    main()
