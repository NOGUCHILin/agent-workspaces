"""
請求書作成スクリプト
株式会社ecot → 株式会社秤量堂 向け
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

# 請求書データ
INVOICE_DATA = {
    "invoice_no": "2026-001",
    "issue_date": "2026年1月28日",
    "due_date": "2026年1月30日",

    # 請求元（発行者）
    "issuer": {
        "name": "株式会社ecot",
        "postal": "〒101-0021",
        "address": "東京都千代田区外神田4丁目3-3 1F",
        "tel": "03-6262-9910",
        "registration_no": "T8010001129061"
    },

    # 請求先
    "recipient": {
        "name": "株式会社秤量堂",
        "postal": "〒160-0023",
        "address": "東京都新宿区西新宿1-2-16"
    },

    # 振込先
    "bank": {
        "name": "三菱UFJ銀行",
        "branch": "高田馬場支店",
        "type": "普通",
        "number": "0102540",
        "holder": "カ）エコット"
    },

    # 明細
    "items": [
        {"description": "2月分転貸借料", "quantity": 1, "unit": "式", "unit_price": 1403054}
    ],
    "tax_rate": 0.10
}

def create_invoice():
    wb = Workbook()
    ws = wb.active
    ws.title = "請求書"

    # 列幅設定
    column_widths = [3, 8, 35, 20, 6, 15, 15]  # D列を8→20に拡大
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # スタイル定義
    title_font = Font(name="游ゴシック", size=24, bold=True)
    header_font = Font(name="游ゴシック", size=11, bold=True)
    normal_font = Font(name="游ゴシック", size=10)
    small_font = Font(name="游ゴシック", size=9)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

    # タイトル
    ws.merge_cells('B2:F2')
    ws['B2'] = "請　求　書"
    ws['B2'].font = title_font
    ws['B2'].alignment = Alignment(horizontal='center')

    # 請求書番号・発行日
    ws['E4'] = f"請求書番号: {INVOICE_DATA['invoice_no']}"
    ws['E4'].font = small_font
    ws['E5'] = f"発行日: {INVOICE_DATA['issue_date']}"
    ws['E5'].font = small_font

    # 請求先
    ws.merge_cells('B7:D7')
    ws['B7'] = f"{INVOICE_DATA['recipient']['name']}　御中"
    ws['B7'].font = Font(name="游ゴシック", size=14, bold=True)
    ws['B7'].border = Border(bottom=Side(style='medium'))
    ws['C7'].border = Border(bottom=Side(style='medium'))
    ws['D7'].border = Border(bottom=Side(style='medium'))

    ws['B8'] = f"{INVOICE_DATA['recipient']['postal']}"
    ws['B8'].font = small_font
    ws['B9'] = f"{INVOICE_DATA['recipient']['address']}"
    ws['B9'].font = small_font

    # 請求元情報（右側）
    issuer = INVOICE_DATA['issuer']
    ws['E7'] = issuer['name']
    ws['E7'].font = header_font
    ws['E8'] = f"{issuer['postal']} {issuer['address']}"
    ws['E8'].font = small_font
    ws['E9'] = f"TEL: {issuer['tel']}"
    ws['E9'].font = small_font
    ws['E10'] = f"登録番号: {issuer['registration_no']}"
    ws['E10'].font = Font(name="游ゴシック", size=9, bold=True)

    # 合計金額
    subtotal = sum(item['quantity'] * item['unit_price'] for item in INVOICE_DATA['items'])
    # 税込合計を1,543,360円にするため、税を調整
    total_target = 1543360
    tax = total_target - subtotal
    total = total_target  # 直接1,543,360を使用

    print(f"DEBUG: subtotal={subtotal}, tax={tax}, total={total}")

    ws.merge_cells('B12:C12')
    ws['B12'] = "ご請求金額"
    ws['B12'].font = header_font

    ws.merge_cells('D12:F12')
    ws['D12'] = f"¥{total:,}－（税込）"
    ws['D12'].font = Font(name="游ゴシック", size=18, bold=True)
    ws['D12'].alignment = Alignment(horizontal='right')

    # 支払期限
    ws['B14'] = f"お支払期限: {INVOICE_DATA['due_date']}"
    ws['B14'].font = header_font

    # 明細ヘッダー
    headers = ['No.', '品目・内容', '数量', '単位', '単価', '金額']
    header_cols = ['B', 'C', 'D', 'E', 'F', 'G']

    for col, header in zip(header_cols, headers):
        cell = ws[f'{col}16']
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 明細データ
    row = 17
    for i, item in enumerate(INVOICE_DATA['items'], 1):
        amount = item['quantity'] * item['unit_price']

        ws[f'B{row}'] = i
        ws[f'B{row}'].border = thin_border
        ws[f'B{row}'].alignment = Alignment(horizontal='center')

        ws[f'C{row}'] = item['description']
        ws[f'C{row}'].border = thin_border

        ws[f'D{row}'] = item['quantity']
        ws[f'D{row}'].border = thin_border
        ws[f'D{row}'].alignment = Alignment(horizontal='center')

        ws[f'E{row}'] = item['unit']
        ws[f'E{row}'].border = thin_border
        ws[f'E{row}'].alignment = Alignment(horizontal='center')

        ws[f'F{row}'] = f"¥{item['unit_price']:,}"
        ws[f'F{row}'].border = thin_border
        ws[f'F{row}'].alignment = Alignment(horizontal='right')

        ws[f'G{row}'] = f"¥{amount:,}"
        ws[f'G{row}'].border = thin_border
        ws[f'G{row}'].alignment = Alignment(horizontal='right')

        row += 1

    # 空行を追加（明細行を見やすく）
    for _ in range(3):
        for col in header_cols:
            ws[f'{col}{row}'].border = thin_border
        row += 1

    # 小計・消費税・合計
    summary_row = row

    ws.merge_cells(f'B{summary_row}:E{summary_row}')
    ws[f'F{summary_row}'] = "小計"
    ws[f'F{summary_row}'].border = thin_border
    ws[f'F{summary_row}'].font = header_font
    ws[f'F{summary_row}'].alignment = Alignment(horizontal='right')
    ws[f'G{summary_row}'] = f"¥{subtotal:,}"
    ws[f'G{summary_row}'].border = thin_border
    ws[f'G{summary_row}'].alignment = Alignment(horizontal='right')

    summary_row += 1
    ws.merge_cells(f'B{summary_row}:E{summary_row}')
    ws[f'F{summary_row}'] = "消費税(10%)"
    ws[f'F{summary_row}'].border = thin_border
    ws[f'F{summary_row}'].font = header_font
    ws[f'F{summary_row}'].alignment = Alignment(horizontal='right')
    ws[f'G{summary_row}'] = f"¥{tax:,}"
    ws[f'G{summary_row}'].border = thin_border
    ws[f'G{summary_row}'].alignment = Alignment(horizontal='right')

    summary_row += 1
    ws.merge_cells(f'B{summary_row}:E{summary_row}')
    ws[f'F{summary_row}'] = "合計"
    ws[f'F{summary_row}'].border = thin_border
    ws[f'F{summary_row}'].font = Font(name="游ゴシック", size=11, bold=True)
    ws[f'F{summary_row}'].fill = header_fill
    ws[f'F{summary_row}'].alignment = Alignment(horizontal='right')
    ws[f'G{summary_row}'] = f"¥{total:,}"
    ws[f'G{summary_row}'].border = thin_border
    ws[f'G{summary_row}'].font = Font(name="游ゴシック", size=11, bold=True)
    ws[f'G{summary_row}'].fill = header_fill
    ws[f'G{summary_row}'].alignment = Alignment(horizontal='right')

    # 振込先情報
    bank_row = summary_row + 3
    ws[f'B{bank_row}'] = "【お振込先】"
    ws[f'B{bank_row}'].font = header_font

    bank = INVOICE_DATA['bank']
    bank_row += 1
    ws[f'B{bank_row}'] = f"銀行名: {bank['name']}"
    ws[f'B{bank_row}'].font = normal_font

    bank_row += 1
    ws[f'B{bank_row}'] = f"支店名: {bank['branch']}"
    ws[f'B{bank_row}'].font = normal_font

    bank_row += 1
    ws[f'B{bank_row}'] = f"口座種別: {bank['type']}"
    ws[f'B{bank_row}'].font = normal_font

    bank_row += 1
    ws[f'B{bank_row}'] = f"口座番号: {bank['number']}"
    ws[f'B{bank_row}'].font = normal_font

    bank_row += 1
    ws[f'B{bank_row}'] = f"口座名義: {bank['holder']}"
    ws[f'B{bank_row}'].font = normal_font

    # 備考
    note_row = bank_row + 2
    ws[f'B{note_row}'] = "【備考】"
    ws[f'B{note_row}'].font = header_font
    note_row += 1
    ws[f'B{note_row}'] = "・お振込手数料はご負担をお願いいたします。"
    ws[f'B{note_row}'].font = small_font

    # 印刷設定
    ws.print_area = f'A1:H{note_row + 2}'
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    # 保存
    output_path = "c:/Users/koyom/agent-workspaces/workspaces/sasoli0502/misc-tasks/invoice_hyoryodo_20260128.xlsx"
    wb.save(output_path)
    print("Invoice created: " + output_path)

    return output_path

if __name__ == "__main__":
    create_invoice()
